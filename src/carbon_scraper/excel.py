"""Builds the deliverable spreadsheet.

Column order comes from `assets/fields-asked.txt` at runtime — that file is the
schema. Adding or reordering a column there changes the output with no code
change.

Every registry lands in one sheet, stacked. The `Standard` column is what
separates them, so the business can filter rather than reconcile two files.

Three extra columns are appended after the requested ones: `Registry` and
`Status` (needed to filter the sheet, since the scrape covers every project
status) and `Project URL` (so any row can be checked against its registry in
one click).

**A delivery is never overwritten.** Each export writes the next version —
`carbon-projects_v1.xlsx`, `_v2`, ... — so a spreadsheet already sent to the
business stays exactly as it was sent. See `next_version_path`.
"""

from __future__ import annotations

import logging
import re
import shutil
import sqlite3
from collections.abc import Sequence
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db, settings

log = logging.getLogger(__name__)

EXTRA_COLUMNS = ["Registry", "Status", "Project URL"]

# Requested column -> projects table column, for the straight copies.
# Every registry's adapter normalises onto these names, so one map serves all.
DIRECT_COLUMNS: dict[str, str] = {
    "Project ID": "external_id",
    "Project Name": "project_name",
    "Standard": "standard_name",
    "Tipo Macro de Projeto": "sectoral_scope",
    "Metodologia": "methodologies",
    "País": "country_name",
    "Estado": "state_province",
    "Cidade": "city",
    "Yearly Ex Ante": "avg_annual_vol_vcu",
    "Status": "status",
}

DATE_COLUMNS: dict[str, tuple[str, ...]] = {
    "Data de Início": ("credit_period_start", "project_start_date"),
    "Data de Término": ("credit_period_end", "project_end_date"),
}

NUMERIC_COLUMNS = {
    "Yearly Ex Ante",
    "Total Ex Ante",
    "Total Credits Issued",
    "Total Credits Sold",
    "Total Credits Retired",
    "Total Credits Cancelled",
    "Duração",
}


def load_credits_config() -> dict[str, Any]:
    path = settings.credits_config()
    if not path.exists():
        return {"sold_equals_retired": True}
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def _to_date(value: Any) -> datetime | None:
    if not value:
        return None
    text = str(value).replace("Z", "")
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _project_url(record: dict[str, Any]) -> str | None:
    """Prefer the URL the adapter recorded; fall back to the registry pattern.

    A registry with no template registered gets a blank cell. It must never
    inherit another registry's pattern: a link that resolves to the wrong
    registry's project page is worse than no link, because it looks right.
    """
    if record.get("detail_url"):
        return str(record["detail_url"])
    project_id = record.get("project_id")
    if project_id is None:
        return None
    template = settings.PROJECT_DETAIL_URLS.get(str(record.get("registry")))
    if not template:
        return None
    return template.format(project_id=project_id)


def build_rows(
    conn: sqlite3.Connection, registry: str | Sequence[str] | None = None
) -> tuple[list[str], list[dict[str, Any]]]:
    """Assemble one dict per project, keyed by the requested column names.

    `registry` is `None` for every registry, a name, or the sequence the GUI's
    checkboxes produce. See `db.registry_clause` for why an empty sequence
    means no rows rather than all of them.
    """
    columns = settings.read_requested_fields()
    for extra in EXTRA_COLUMNS:
        if extra not in columns:
            columns.append(extra)

    config = load_credits_config()
    sold_equals_retired = bool(config.get("sold_equals_retired", True))
    totals_map = config.get("totals") or {
        "Total Credits Issued": "issuances",
        "Total Credits Retired": "retirements",
        "Total Credits Cancelled": "cancellations",
    }

    totals = db.credit_totals(conn)
    by_beneficiary = db.retired_by_beneficiary(conn)
    certifications = db.certifications_by_project(conn)
    derived = db.derived_map(conn)

    rows: list[dict[str, Any]] = []
    for project in db.all_projects(conn, registry):
        record = dict(project)
        key = (record["registry"], record["project_id"])
        project_totals = totals.get(key, {})
        project_derived = derived.get(key, {})

        row: dict[str, Any] = {}
        for column in columns:
            if column == "Registry":
                row[column] = settings.REGISTRY_LABELS.get(
                    record["registry"], record["registry"]
                )
            elif column == "Standard":
                # Gold Standard states its standard version; Verra states the
                # standard name. Either way, fall back to the registry label
                # rather than leaving the discriminator column empty.
                row[column] = record.get("standard_name") or settings.REGISTRY_LABELS.get(
                    record["registry"], record["registry"]
                )
            elif column == "Project ID":
                row[column] = record.get("external_id") or record.get("project_id")
            elif column == "Continent":
                # Verra publishes a region; Gold Standard does not, so its
                # rows fall through to the country-code rules in
                # config/derivation/continent.yaml.
                row[column] = record.get("region_name") or project_derived.get(column)
            elif column in DIRECT_COLUMNS:
                row[column] = record.get(DIRECT_COLUMNS[column])
            elif column in DATE_COLUMNS:
                row[column] = next(
                    (d for d in (_to_date(record.get(f)) for f in DATE_COLUMNS[column]) if d),
                    None,
                )
            elif column == "Additional Certification":
                row[column] = record.get("additional_certification") or certifications.get(key)
            elif column == "Total Credits Sold":
                row[column] = (
                    project_totals.get("retirements")
                    if sold_equals_retired
                    else by_beneficiary.get(key)
                )
            elif column in totals_map:
                row[column] = project_totals.get(totals_map[column])
            elif column == "Project URL":
                row[column] = _project_url(record)
            else:
                # Derived columns (Tipo Micro, Bioma, Durabilidade, Duração,
                # Total Ex Ante). Blank when no rule matched — never invented.
                row[column] = project_derived.get(column)

            if column in NUMERIC_COLUMNS and row[column] is not None:
                try:
                    number = float(row[column])
                    row[column] = int(number) if number.is_integer() else number
                except (TypeError, ValueError):
                    pass

        rows.append(row)
    return columns, rows


# -- versioned output ------------------------------------------------------

_VERSION_RE = re.compile(r"_v(\d+)$", re.I)

# Deliverables written before the versioning rule existed. The newest of these
# is adopted as v1 so the version history starts from what was actually sent.
LEGACY_NAMES = ("carbon-projects.xlsx", "verra-projects.xlsx")


def existing_versions(base: Path | None = None) -> list[tuple[int, Path]]:
    base = Path(base or settings.XLSX_BASE)
    found: list[tuple[int, Path]] = []
    for path in base.parent.glob(f"{base.stem}_v*{base.suffix}"):
        match = _VERSION_RE.search(path.stem)
        if match:
            found.append((int(match.group(1)), path))
    return sorted(found)


def next_version_path(base: Path | None = None) -> tuple[Path, Path | None]:
    """Return (path_to_write, previous_delivery).

    Never returns a path that already exists: the last delivery is left on
    disk untouched and the new build goes to the next version number. If no
    versioned file exists yet but an older unversioned deliverable does, that
    one is adopted as `_v1` first, so history starts from the sheet the
    business actually received.
    """
    base = Path(base or settings.XLSX_BASE)
    base.parent.mkdir(parents=True, exist_ok=True)
    versions = existing_versions(base)

    if not versions:
        legacy = [base.parent / name for name in LEGACY_NAMES]
        legacy = [p for p in legacy if p.exists()]
        if legacy:
            newest = max(legacy, key=lambda p: p.stat().st_mtime)
            first = base.with_name(f"{base.stem}_v1{base.suffix}")
            shutil.copy2(newest, first)
            log.info("Adopted %s as %s.", newest.name, first.name)
            return base.with_name(f"{base.stem}_v2{base.suffix}"), first
        return base.with_name(f"{base.stem}_v1{base.suffix}"), None

    latest_number, latest_path = versions[-1]
    return (
        base.with_name(f"{base.stem}_v{latest_number + 1}{base.suffix}"),
        latest_path,
    )


def write_xlsx(
    conn: sqlite3.Connection,
    path: Any = None,
    *,
    registry: str | Sequence[str] | None = None,
    out_dir: Any = None,
) -> tuple[Path, int, Path | None]:
    """Write the next version of the spreadsheet.

    Returns (written_path, row_count, previous_delivery).

    `out_dir` chooses the folder while keeping versioning — that is what the
    GUI's folder picker uses. An explicit `path` bypasses versioning entirely
    and exists for tests; it is not a "just overwrite it" flag.
    """
    settings.ensure_dirs()
    previous: Path | None = None
    if path is None:
        base = Path(out_dir) / settings.XLSX_BASE.name if out_dir else None
        path, previous = next_version_path(base)
    path = Path(path)

    columns, rows = build_rows(conn, registry)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Carbon Projects"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=2):
        for col_index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=row.get(column))
            if isinstance(row.get(column), datetime):
                cell.number_format = "yyyy-mm-dd"
            elif column in NUMERIC_COLUMNS and isinstance(row.get(column), (int, float)):
                cell.number_format = "#,##0"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(columns))}{max(len(rows) + 1, 2)}"
    )

    widths = {"Project Name": 46, "Metodologia": 26, "Project URL": 48, "Registry": 16}
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(
            column, max(14, min(len(column) + 4, 34))
        )

    workbook.save(path)
    return path, len(rows), previous


def coverage_report(
    conn: sqlite3.Connection, registry: str | Sequence[str] | None = None
) -> list[tuple[str, int, int]]:
    """(column, filled, total) — how complete each column actually is."""
    columns, rows = build_rows(conn, registry)
    total = len(rows)
    return [
        (column, sum(1 for r in rows if r.get(column) not in (None, "")), total)
        for column in columns
    ]
